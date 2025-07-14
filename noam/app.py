from extractor import extract_inventory_items_prompt
import pandas as pd
import time
from openai import OpenAI
from openpyxl import load_workbook
from colorama import Fore, Back, Style
from io import StringIO
from email_parser import parse_email

client = OpenAI(
    base_url="http://localhost:8080/v1",
    api_key="tom",
)

time.sleep(5)

def prompt(email_body):
    response = client.chat.completions.create(
            model="llama.cpp/models/mMistral-Small-3.2-24B-Instruct-2506-Q4_K_M.gguf",
            messages=[
                {
                    "role": "user",
                    "content": extract_inventory_items_prompt(email_body),
                    # "content": extract_inventory_items_prompt("Hi, could I please get pricing for these items below and how fast you could get them shipped to Huntington Beach location.Alkemix 5-finger Lap Shear Panels in Alclad 2024-T3 quantity 100 Alkemix 5-finger Lap Shear Panels in Bare 7075-T6 quantity 50"),
                }
            ],
            stream=False,
            max_tokens=1000,
            )
                
    output = response.choices[0].message.content
    output = output.removeprefix("```json\n")
    output = output.removesuffix("\n```")
    
    return output

email = parse_email("data/Part request 2.eml")

print(email.body)
txt = prompt(email.body)

print(txt)
df = pd.read_json(StringIO(txt))
print(df.head())

# df = pd.read_json('data/data.json')
# print(df.head())

# df2 = pd.read_json('data/data2.json')
# print(df2.head())

# Load the existing workbook to preserve all formatting
wb = load_workbook('data/Boeing MO.xlsx')
ws = wb['Input Form']  # or whatever your sheet name is

# Update only the specific cells with our data (preserving all formatting)
for i, row_data in enumerate(df.values):
    row_num = 23 + i  # Start at row 22 (since we skipped 21 rows)
    
    # Update cells A, B, C, D for this row
    ws[f'A{row_num}'] = i + 1
    ws[f'B{row_num}'] = row_data[1]  # MFCTR P/N
    ws[f'C{row_num}'] = row_data[0] + ': ' + row_data[3]  # DESCRIPTION
    ws[f'D{row_num}'] = row_data[2]  # QTY

# Save the workbook (this preserves all formatting, dropdowns, colors, etc.)
wb.save('data/Boeing MO Filled.xlsx')

print("Done!")