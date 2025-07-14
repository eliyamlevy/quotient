import sys
import os
import json
import pandas as pd
from pathlib import Path
from typing import Optional, List
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QTextEdit, QFileDialog, QProgressBar,
    QTableWidget, QTableWidgetItem, QTabWidget, QMessageBox,
    QGroupBox, QGridLayout, QSpinBox, QDoubleSpinBox, QComboBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QFont, QIcon, QPixmap
import time
from openai import OpenAI
from io import StringIO
from openpyxl import load_workbook

# Import your existing modules
from email_parser import parse_email
from extractor import extract_inventory_items_prompt

class AIProcessor(QThread):
    """Background thread for AI processing"""
    progress_updated = pyqtSignal(str)
    processing_complete = pyqtSignal(dict)
    error_occurred = pyqtSignal(str)
    debug_log = pyqtSignal(str)
    
    def __init__(self, email_body: str, provider: str, model_name: str, max_tokens: int, 
                 temperature: float, server_url: str = "", api_key: str = ""):
        super().__init__()
        self.email_body = email_body
        self.provider = provider
        self.model_name = model_name
        self.max_tokens = max_tokens
        self.temperature = temperature
        self.server_url = server_url
        self.api_key = api_key
        
    def run(self):
        try:
            self.progress_updated.emit("Initializing AI client...")
            
            if self.provider == "Local AI Server":
                self.debug_log.emit(f"Initializing OpenAI client with local server: {self.server_url}")
                client = OpenAI(
                    base_url=self.server_url,
                    api_key=self.api_key,
                )
            else:
                self.debug_log.emit("Initializing OpenAI client with ChatGPT")
                client = OpenAI(
                    api_key=self.api_key,
                )
            
            self.progress_updated.emit("Sending request to AI model...")
            self.debug_log.emit(f"Sending request to model: {self.model_name}")
            self.debug_log.emit(f"Max tokens: {self.max_tokens}")
            
            response = client.chat.completions.create(
                model=self.model_name,
                messages=[
                    {
                        "role": "user",
                        "content": extract_inventory_items_prompt(self.email_body),
                    }
                ],
                stream=False,
                max_tokens=self.max_tokens,
                temperature=self.temperature,
            )
            
            self.progress_updated.emit("Processing AI response...")
            output = response.choices[0].message.content
            
            # Log the raw LLM output
            debug_output = "="*50 + "\n"
            debug_output += "RAW LLM OUTPUT:\n"
            debug_output += "="*50 + "\n"
            debug_output += output + "\n"
            debug_output += "="*50 + "\n"
            debug_output += "END RAW OUTPUT\n"
            debug_output += "="*50
            self.debug_log.emit(debug_output)
            self.progress_updated.emit(f"Raw LLM Output received (see debug log)")
            output = output.removeprefix("```json\n")
            output = output.removesuffix("\n```")
            
            # Parse the JSON response
            try:
                df = pd.read_json(StringIO(output))
                self.progress_updated.emit("JSON parsing successful!")
            except Exception as json_error:
                error_msg = f"JSON parsing failed: {str(json_error)}\nRaw output was:\n{output}"
                self.debug_log.emit(error_msg)
                self.error_occurred.emit(error_msg)
                return
            
            self.progress_updated.emit("Processing complete!")
            self.processing_complete.emit({
                'raw_response': output,
                'dataframe': df,
                'items_count': len(df)
            })
            
        except Exception as e:
            self.error_occurred.emit(f"Error during processing: {str(e)}")

class EmailProcessorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_data = None
        self.current_email_file = None
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("Email Inventory Processor")
        self.setGeometry(100, 100, 1200, 800)
        
        # Set up the main widget and layout
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        
        # Create tab widget
        tabs = QTabWidget()
        layout.addWidget(tabs)
        
        # Create tabs
        tabs.addTab(self.create_home_tab(), "Home")
        tabs.addTab(self.create_ai_config_tab(), "AI Config")
        tabs.addTab(self.create_processing_tab(), "Process Email")
        tabs.addTab(self.create_results_tab(), "Results")
        tabs.addTab(self.create_raw_output_tab(), "Raw Output")
        tabs.addTab(self.create_debug_tab(), "Debug Log")
        tabs.addTab(self.create_export_tab(), "Export")
        
        # Status bar
        self.status_bar = self.statusBar()
        self.status_bar.showMessage("Ready to process emails")
        
    def create_processing_tab(self):
        """Create the main processing tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # File selection group
        file_group = QGroupBox("Email File Selection")
        file_layout = QVBoxLayout(file_group)
        
        # File selection row
        file_row = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("padding: 5px; border: 1px solid #ccc; background: #f9f9f9;")
        select_file_btn = QPushButton("Select Email File (.eml)")
        select_file_btn.clicked.connect(self.select_email_file)
        file_row.addWidget(self.file_label, 1)
        file_row.addWidget(select_file_btn)
        file_layout.addLayout(file_row)
        
        layout.addWidget(file_group)
        
        # Email preview group
        preview_group = QGroupBox("Email Preview")
        preview_layout = QVBoxLayout(preview_group)
        
        self.email_preview = QTextEdit()
        self.email_preview.setMaximumHeight(200)
        self.email_preview.setReadOnly(True)
        preview_layout.addWidget(self.email_preview)
        
        layout.addWidget(preview_group)
        
        # Processing controls
        process_group = QGroupBox("Processing")
        process_layout = QVBoxLayout(process_group)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        process_layout.addWidget(self.progress_bar)
        
        # Process button
        self.process_btn = QPushButton("Process Email")
        self.process_btn.clicked.connect(self.process_email)
        self.process_btn.setEnabled(False)
        process_layout.addWidget(self.process_btn)
        
        layout.addWidget(process_group)
        
        # Status display
        self.status_display = QTextEdit()
        self.status_display.setMaximumHeight(100)
        self.status_display.setReadOnly(True)
        layout.addWidget(self.status_display)
        
        layout.addStretch()
        return widget
    
    def create_home_tab(self):
        """Create the home/welcome tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Center the content
        layout.addStretch()
        
        # App title
        title_label = QLabel("Quotient")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_font = QFont()
        title_font.setPointSize(50)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setStyleSheet("color: white; margin: 20px;")
        layout.addWidget(title_label)
        
        # App description
        desc_label = QLabel(
            "Email Inventory Processor\n\n"
            "Quotient is an intelligent email processing application that extracts inventory items "
            "from email content using AI. Simply load an email file, configure your AI settings, "
            "and let the system automatically identify and extract part numbers, quantities, and descriptions. "
            "Export the results to Excel templates or JSON format for further processing."
        )
        desc_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc_label.setWordWrap(True)
        desc_font = QFont()
        desc_font.setPointSize(36)
        desc_label.setFont(desc_font)
        desc_label.setStyleSheet("color: white; margin: 20px; line-height: 1.5;")
        layout.addWidget(desc_label)
        
        # Quick start section
        quick_start_group = QGroupBox("Quick Start")
        quick_start_layout = QVBoxLayout(quick_start_group)
        
        quick_start_text = QLabel(
            "1. Go to 'Process Email' tab\n"
            "2. Select an email file (.eml)\n"
            "3. Configure AI settings (optional)\n"
            "4. Click 'Process Email' to extract items\n"
            "5. View results and export as needed"
        )
        quick_start_text.setStyleSheet("color: #7f8c8d; margin: 10px;")
        quick_start_layout.addWidget(quick_start_text)
        
        layout.addWidget(quick_start_group)
        
        layout.addStretch()
        return widget
    
    def create_ai_config_tab(self):
        """Create the AI configuration tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # AI Provider Selection
        provider_group = QGroupBox("AI Provider")
        provider_layout = QVBoxLayout(provider_group)
        
        self.provider_combo = QComboBox()
        self.provider_combo.addItems(["Local AI Server", "ChatGPT (OpenAI)"])
        self.provider_combo.currentTextChanged.connect(self.on_provider_changed)
        provider_layout.addWidget(self.provider_combo)
        
        layout.addWidget(provider_group)
        
        # Local AI Settings
        self.local_ai_group = QGroupBox("Local AI Server Settings")
        local_layout = QGridLayout(self.local_ai_group)
        
        # Model selection
        local_layout.addWidget(QLabel("AI Model:"), 0, 0)
        self.model_combo = QComboBox()
        self.model_combo.addItems([
            "llama.cpp/models/mMistral-Small-3.2-24B-Instruct-2506-Q4_K_M.gguf",
            "llama.cpp/models/llama-2-7b-chat.gguf",
            "llama.cpp/models/codellama-7b-instruct.gguf"
        ])
        local_layout.addWidget(self.model_combo, 0, 1)
        
        # Server URL
        local_layout.addWidget(QLabel("Server URL:"), 1, 0)
        self.server_url_input = QTextEdit()
        self.server_url_input.setMaximumHeight(30)
        self.server_url_input.setPlainText("http://localhost:8080/v1")
        local_layout.addWidget(self.server_url_input, 1, 1)
        
        # API Key
        local_layout.addWidget(QLabel("API Key:"), 2, 0)
        self.local_api_key_input = QTextEdit()
        self.local_api_key_input.setMaximumHeight(30)
        self.local_api_key_input.setPlainText("tom")
        local_layout.addWidget(self.local_api_key_input, 2, 1)
        
        layout.addWidget(self.local_ai_group)
        
        # ChatGPT Settings
        self.chatgpt_group = QGroupBox("ChatGPT Settings")
        chatgpt_layout = QGridLayout(self.chatgpt_group)
        
        # ChatGPT API Key
        chatgpt_layout.addWidget(QLabel("OpenAI API Key:"), 0, 0)
        self.chatgpt_api_key_input = QTextEdit()
        self.chatgpt_api_key_input.setMaximumHeight(30)
        self.chatgpt_api_key_input.setPlaceholderText("Enter your OpenAI API key (sk-...)")
        chatgpt_layout.addWidget(self.chatgpt_api_key_input, 0, 1)
        
        # ChatGPT Model
        chatgpt_layout.addWidget(QLabel("Model:"), 1, 0)
        self.chatgpt_model_combo = QComboBox()
        self.chatgpt_model_combo.addItems([
            "gpt-4",
            "gpt-4-turbo-preview",
            "gpt-3.5-turbo",
            "gpt-3.5-turbo-16k"
        ])
        chatgpt_layout.addWidget(self.chatgpt_model_combo, 1, 1)
        
        layout.addWidget(self.chatgpt_group)
        
        # Common Settings
        common_group = QGroupBox("Common Settings")
        common_layout = QGridLayout(common_group)
        
        # Max tokens
        common_layout.addWidget(QLabel("Max Tokens:"), 0, 0)
        self.max_tokens_spin = QSpinBox()
        self.max_tokens_spin.setRange(100, 4000)
        self.max_tokens_spin.setValue(1000)
        common_layout.addWidget(self.max_tokens_spin, 0, 1)
        
        # Temperature
        common_layout.addWidget(QLabel("Temperature:"), 1, 0)
        self.temperature_spin = QDoubleSpinBox()
        self.temperature_spin.setRange(0.0, 2.0)
        self.temperature_spin.setValue(0.1)
        self.temperature_spin.setSingleStep(0.1)
        common_layout.addWidget(self.temperature_spin, 1, 1)
        
        layout.addWidget(common_group)
        
        # Test Connection Button
        test_btn = QPushButton("Test Connection")
        test_btn.clicked.connect(self.test_ai_connection)
        layout.addWidget(test_btn)
        
        layout.addStretch()
        return widget
    
    def on_provider_changed(self, provider):
        """Handle AI provider selection change"""
        if provider == "Local AI Server":
            self.local_ai_group.setVisible(True)
            self.chatgpt_group.setVisible(False)
        else:
            self.local_ai_group.setVisible(False)
            self.chatgpt_group.setVisible(True)
    
    def test_ai_connection(self):
        """Test the AI connection"""
        try:
            provider = self.provider_combo.currentText()
            
            if provider == "Local AI Server":
                client = OpenAI(
                    base_url=self.server_url_input.toPlainText().strip(),
                    api_key=self.local_api_key_input.toPlainText().strip(),
                )
            else:
                client = OpenAI(
                    api_key=self.chatgpt_api_key_input.toPlainText().strip(),
                )
            
            # Simple test request
            response = client.chat.completions.create(
                model="test" if provider == "Local AI Server" else "gpt-3.5-turbo",
                messages=[{"role": "user", "content": "Hello"}],
                max_tokens=5,
            )
            
            QMessageBox.information(self, "Success", "AI connection test successful!")
            
        except Exception as e:
            QMessageBox.critical(self, "Connection Error", f"Failed to connect: {str(e)}")
    
    def create_results_tab(self):
        """Create the results display tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Results info
        info_group = QGroupBox("Processing Results")
        info_layout = QHBoxLayout(info_group)
        
        self.items_count_label = QLabel("Items extracted: 0")
        self.processing_time_label = QLabel("Processing time: --")
        info_layout.addWidget(self.items_count_label)
        info_layout.addWidget(self.processing_time_label)
        info_layout.addStretch()
        
        layout.addWidget(info_group)
        
        # Results table
        self.results_table = QTableWidget()
        layout.addWidget(self.results_table)
        
        return widget
    
    def create_raw_output_tab(self):
        """Create the raw output display tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Raw output display
        output_group = QGroupBox("Raw LLM Output")
        output_layout = QVBoxLayout(output_group)
        
        self.raw_output_display = QTextEdit()
        self.raw_output_display.setReadOnly(True)
        self.raw_output_display.setFont(QFont("Courier", 10))
        output_layout.addWidget(self.raw_output_display)
        
        layout.addWidget(output_group)
        
        return widget
    
    def create_debug_tab(self):
        """Create the debug log tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Debug log display
        debug_group = QGroupBox("Debug Log")
        debug_layout = QVBoxLayout(debug_group)
        
        self.debug_log = QTextEdit()
        self.debug_log.setReadOnly(True)
        self.debug_log.setFont(QFont("Courier", 9))
        self.debug_log.setStyleSheet("background-color: #1e1e1e; color: #ffffff;")
        debug_layout.addWidget(self.debug_log)
        
        # Control buttons
        button_layout = QHBoxLayout()
        
        clear_btn = QPushButton("Clear Log")
        clear_btn.clicked.connect(self.clear_debug_log)
        button_layout.addWidget(clear_btn)
        
        save_btn = QPushButton("Save Log")
        save_btn.clicked.connect(self.save_debug_log)
        button_layout.addWidget(save_btn)
        
        button_layout.addStretch()
        debug_layout.addLayout(button_layout)
        
        layout.addWidget(debug_group)
        return widget
    
    def log_debug(self, message):
        """Add a message to the debug log"""
        timestamp = time.strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.debug_log.append(log_entry)
        # Auto-scroll to bottom
        self.debug_log.verticalScrollBar().setValue(
            self.debug_log.verticalScrollBar().maximum()
        )
    
    def clear_debug_log(self):
        """Clear the debug log"""
        self.debug_log.clear()
    
    def save_debug_log(self):
        """Save the debug log to a file"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Debug Log", "", "Text Files (*.txt);;All Files (*)"
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(self.debug_log.toPlainText())
                QMessageBox.information(self, "Success", f"Debug log saved to: {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save log: {str(e)}")
    
    def create_export_tab(self):
        """Create the export tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Export options
        export_group = QGroupBox("Export Options")
        export_layout = QVBoxLayout(export_group)
        
        # Excel export
        excel_row = QHBoxLayout()
        excel_row.addWidget(QLabel("Excel Template:"))
        self.template_label = QLabel("No template selected")
        self.template_label.setStyleSheet("padding: 5px; border: 1px solid #ccc; background: #f9f9f9;")
        select_template_btn = QPushButton("Select Template")
        select_template_btn.clicked.connect(self.select_excel_template)
        excel_row.addWidget(self.template_label, 1)
        excel_row.addWidget(select_template_btn)
        export_layout.addLayout(excel_row)
        
        # Output filename
        filename_row = QHBoxLayout()
        filename_row.addWidget(QLabel("Output Filename:"))
        self.filename_input = QTextEdit()
        self.filename_input.setMaximumHeight(30)
        self.filename_input.setPlaceholderText("Leave empty for auto-generated name (e.g., Processed_1234567890.xlsx)")
        filename_row.addWidget(self.filename_input, 1)
        export_layout.addLayout(filename_row)
        
        # Export button
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        export_layout.addWidget(self.export_btn)
        
        layout.addWidget(export_group)
        
        # JSON export
        json_group = QGroupBox("JSON Export")
        json_layout = QVBoxLayout(json_group)
        
        json_export_btn = QPushButton("Export as JSON")
        json_export_btn.clicked.connect(self.export_to_json)
        json_layout.addWidget(json_export_btn)
        
        layout.addWidget(json_group)
        
        layout.addStretch()
        return widget
    
    def select_email_file(self):
        """Select an email file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Email File", "", "Email Files (*.eml);;All Files (*)"
        )
        
        if file_path:
            self.current_email_file = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.process_btn.setEnabled(True)
            self.load_email_preview(file_path)
    
    def load_email_preview(self, file_path: str):
        """Load and display email preview"""
        try:
            email = parse_email(file_path)
            preview_text = f"From: {email.header.get('From', 'Unknown')}\n"
            preview_text += f"Subject: {email.header.get('Subject', 'No Subject')}\n"
            preview_text += f"Date: {email.header.get('Date', 'Unknown')}\n"
            preview_text += f"\n{'='*50}\n\n"
            preview_text += email.body[:500] + "..." if len(email.body) > 500 else email.body
            
            self.email_preview.setText(preview_text)
            self.status_bar.showMessage(f"Loaded email: {os.path.basename(file_path)}")
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load email: {str(e)}")
    
    def select_excel_template(self):
        """Select Excel template file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel Template", "", "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            self.template_label.setText(os.path.basename(file_path))
            self.export_btn.setEnabled(self.current_data is not None)
    
    def process_email(self):
        """Process the selected email"""
        if not self.current_email_file:
            return
        
        try:
            # Parse email
            email = parse_email(self.current_email_file)
            
            # Get AI configuration
            provider = self.provider_combo.currentText()
            
            if provider == "Local AI Server":
                model_name = self.model_combo.currentText()
                server_url = self.server_url_input.toPlainText().strip()
                api_key = self.local_api_key_input.toPlainText().strip()
            else:
                model_name = self.chatgpt_model_combo.currentText()
                server_url = ""
                api_key = self.chatgpt_api_key_input.toPlainText().strip()
            
            # Start AI processing in background
            self.processor = AIProcessor(
                email.body,
                provider,
                model_name,
                self.max_tokens_spin.value(),
                self.temperature_spin.value(),
                server_url,
                api_key
            )
            
            self.processor.progress_updated.connect(self.update_progress)
            self.processor.processing_complete.connect(self.on_processing_complete)
            self.processor.error_occurred.connect(self.on_processing_error)
            self.processor.debug_log.connect(self.log_debug)
            
            # Update UI
            self.process_btn.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)  # Indeterminate progress
            self.status_display.clear()
            self.status_display.append("Starting processing...")
            
            # Start processing
            self.processor.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to start processing: {str(e)}")
    
    def update_progress(self, message: str):
        """Update progress display"""
        self.status_display.append(message)
        self.status_bar.showMessage(message)
    
    def on_processing_complete(self, result: dict):
        """Handle processing completion"""
        self.current_data = result
        self.progress_bar.setVisible(False)
        self.process_btn.setEnabled(True)
        
        # Update results tab
        self.display_results(result['dataframe'])
        self.items_count_label.setText(f"Items extracted: {result['items_count']}")
        
        # Update raw output tab
        self.raw_output_display.setText(result['raw_response'])
        
        # Enable export
        self.export_btn.setEnabled(True)
        
        # Clear filename input for new export
        self.filename_input.clear()
        
        # Switch to results tab
        self.centralWidget().findChild(QTabWidget).setCurrentIndex(1)
        
        QMessageBox.information(self, "Success", "Email processing completed successfully!")
    
    def on_processing_error(self, error: str):
        """Handle processing error"""
        self.progress_bar.setVisible(False)
        self.process_btn.setEnabled(True)
        QMessageBox.critical(self, "Processing Error", error)
    
    def display_results(self, df: pd.DataFrame):
        """Display results in table"""
        self.results_table.setRowCount(len(df))
        self.results_table.setColumnCount(len(df.columns))
        self.results_table.setHorizontalHeaderLabels(df.columns)
        
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.results_table.setItem(i, j, item)
        
        self.results_table.resizeColumnsToContents()
    
    def export_to_excel(self):
        """Export results to Excel"""
        if not self.current_data or not self.template_label.text() != "No template selected":
            QMessageBox.warning(self, "Warning", "Please select a template and ensure data is processed")
            return
        
        try:
            template_path = self.template_label.text()
            if not os.path.isabs(template_path):
                # Assume it's in the data directory
                template_path = os.path.join("data", template_path)
            
            # Load template
            wb = load_workbook(template_path)
            ws = wb['Input Form']  # Adjust sheet name as needed
            
            df = self.current_data['dataframe']
            
            # Update cells
            for i, row_data in enumerate(df.values):
                row_num = 23 + i  # Adjust starting row as needed
                
                ws[f'A{row_num}'] = i + 1
                ws[f'B{row_num}'] = row_data[1] if len(row_data) > 1 else ""  # MFCTR P/N
                ws[f'C{row_num}'] = f"{row_data[0]}: {row_data[3]}" if len(row_data) > 3 else str(row_data[0])  # DESCRIPTION
                ws[f'D{row_num}'] = row_data[2] if len(row_data) > 2 else ""  # QTY
            
            # Get custom filename or use default
            custom_filename = self.filename_input.toPlainText().strip()
            if custom_filename:
                # Clean the filename - remove invalid characters
                import re
                custom_filename = re.sub(r'[<>:"/\\|?*]', '_', custom_filename)
                
                # Ensure it has .xlsx extension
                if not custom_filename.lower().endswith('.xlsx'):
                    custom_filename += '.xlsx'
                
                # Check if filename is not empty after cleaning
                if custom_filename.strip('.xlsx').strip():
                    output_path = os.path.join("data", custom_filename)
                else:
                    # Use auto-generated name if cleaned filename is empty
                    output_path = os.path.join("data", f"Processed_{int(time.time())}.xlsx")
            else:
                # Use auto-generated name
                output_path = os.path.join("data", f"Processed_{int(time.time())}.xlsx")
            
            wb.save(output_path)
            
            QMessageBox.information(self, "Success", f"Exported to: {output_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")
    
    def export_to_json(self):
        """Export results to JSON"""
        if not self.current_data:
            QMessageBox.warning(self, "Warning", "No data to export")
            return
        
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save JSON", "", "JSON Files (*.json)"
            )
            
            if file_path:
                df = self.current_data['dataframe']
                df.to_json(file_path, orient='records', indent=2)
                QMessageBox.information(self, "Success", f"Exported to: {file_path}")
                
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

def main():
    app = QApplication(sys.argv)
    
    # Set application style
    app.setStyle('Fusion')
    
    # Create and show the main window
    window = EmailProcessorGUI()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main() 